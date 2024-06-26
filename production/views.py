from rest_framework import generics
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.pagination import PageNumberPagination
from rest_framework.parsers import JSONParser
from django.http.response import JsonResponse
from django.db import connection, OperationalError
from rest_framework.decorators import action
from rest_framework import viewsets, serializers, status
from datetime import timedelta, datetime
from rest_framework.response import Response
from django.db.models import Count
from datetime import datetime, time, date
from django.db.models import Q, Count, Sum, F, FloatField

from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter, column_index_from_string
from django.views.generic import View
from django.http import HttpResponse
from rest_framework.parsers import FileUploadParser
from rest_framework.views import APIView
import pandas as pd
from pytz import timezone
from openpyxl.utils import get_column_letter



from production import models
from production import serializers

from config import models as config_models



class defaultPagination(PageNumberPagination):
    page_size = 10
    page_size_query_param = 'page_size'
    max_page_size = 100


class productionPlanningList(generics.ListCreateAPIView):
    queryset = models.productionPlanning.objects.order_by('-id')
    serializer_class = serializers.productionPlanningSerializer
    pagination_class = defaultPagination

class MostOrderedProducts(generics.RetrieveAPIView):
    def get(self, request, *args, **kwargs):
        total_rows = models.productionPlanning.objects.count()
        product_counts = models.productionPlanning.objects.values('product_id').annotate(count=Count('product_id'))
        most_ordered_products = []

        for entry in product_counts:
            product_id = entry['product_id']
            count = entry['count']
            percentage = (count / total_rows) * 100

            most_ordered_products.append({
                'product_id': product_id,
                'category': product_id,
                'percentage': percentage
            })

        return Response({'most_ordered_products': most_ordered_products})
    
class recentOrders(generics.ListAPIView):
    queryset = models.productionPlanning.objects.order_by('-id')
    serializer_class = serializers.recentOrdersSerializer
    pagination_class = defaultPagination

class productionPlanningEdit(generics.RetrieveUpdateDestroyAPIView):
    queryset = models.productionPlanning.objects.all()
    serializer_class = serializers.productionPlanningSerializer
    lookup_url_kwarg = "id"

class lineMachineConfigList(generics.ListCreateAPIView):
    queryset = models.lineMachineConfig.objects.order_by('-id')
    serializer_class = serializers.lineMachineConfigSerializer
    pagination_class = defaultPagination

class lineMachineConfigEdit(generics.RetrieveUpdateDestroyAPIView):
    queryset = models.lineMachineConfig.objects.all()
    serializer_class = serializers.lineMachineConfigSerializer
    lookup_url_kwarg = "id"



# Database Connection Check Function
def check_database_connection(request):
    try:
        with connection.cursor():
            # Try to execute a simple query to check the database connection
            pass
        return JsonResponse({'message': 'Database Connected.'})
    except OperationalError as e:
        # Handle the database connection error
        error_message = str(e)
        return JsonResponse({'error': 'Database connection error', 'message': error_message}, status=500)
    


# Open Jobworks from Production Planning
class openJobWorks(generics.ListCreateAPIView):
    # queryset = models.productionPlanning.objects.raw('SELECT id, job_id FROM production_planning WHERE planned_date IS NULL ORDER BY id ASC').no_cache()
    queryset = models.productionPlanning.objects.filter(planned_date__isnull=True).order_by('id')
    serializer_class = serializers.openJobWorkSerializer

class productionPlanById(generics.ListCreateAPIView):
    queryset = models.productionPlanning.objects.all()
    serializer_class = serializers.productionPlanningSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['id', 'job_id']
    

# class LineMachineSlotConfigViewSet(viewsets.ModelViewSet):
#     queryset = models.lineMachineSlotConfig.objects.all()
#     serializer_class = serializers.lineMachineSlotConfigSerializer

#     @action(detail=False, methods=['POST'])
#     def calculate_schedule(self, request):
#         # Get input data from the request
#         product_id = request.data.get('product_id')
#         quantity = request.data.get('quantity')
#         start_date = datetime.strptime(request.data.get('start_date'), '%Y-%m-%d')
        
#         # Define the product target (60 seconds per unit)
#         product_target = 60  # seconds
        
#         # Calculate total hours required to produce the quantity
#         total_seconds_required = quantity * product_target
#         total_hours_required = total_seconds_required / 3600  # 1 hour = 3600 seconds

#         # Calculate the number of days required
#         days_required = total_hours_required / 21  # Assuming 21 hours per day

#         # Initialize the current date and remaining hours
#         current_date = start_date
#         remaining_hours = total_hours_required
#         planned_production = 0

#         while remaining_hours > 0:
#             # Calculate planned production for the day
#             planned_production = min(21, remaining_hours)
            
#             # Create a LineMachineSlotConfig object and save it to the database
#             line_config = models.lineMachineSlotConfig(
#                 product_id=product_id,
#                 date=current_date,
#                 planned_hours=planned_production,
#                 remaining_hours=remaining_hours - planned_production,
#                 balance_production=quantity - planned_production,
#                 shift_a=f'Shift A {planned_production}',
#                 shift_b=f'Shift B {planned_production}',
#                 shift_c=f'Shift C {planned_production}',
#             )
#             line_config.save()
            
#             # Update current date and remaining hours for the next iteration
#             current_date += timedelta(days=1)
#             remaining_hours -= planned_production

#         return Response({'message': 'Schedule calculated and saved.'})





'''Data Processing without consideration of start date and start time
    perfectly working API'''
# class LineMachineSlotConfigViewSet(viewsets.ModelViewSet):
#     queryset = models.lineMachineSlotConfig.objects.all()
#     serializer_class = serializers.lineMachineSlotConfigSerializer
#     filter_backends = [DjangoFilterBackend]
#     filterset_fields = ['id', 'job_id']

#     @action(detail=False, methods=['POST'])
#     def calculate_schedule(self, request):
#         # Deserialize input data
#         input_serializer = serializers.ScheduleInputSerializer(data=request.data)
#         if input_serializer.is_valid():
#             product_id = input_serializer.validated_data['product_id']
#             quantity = input_serializer.validated_data['quantity']
#             start_date = input_serializer.validated_data['start_date']
#             job_id = input_serializer.validated_data['job_id']
#             company = input_serializer.validated_data['company']
#             plant = input_serializer.validated_data['plant']
#             shopfloor = input_serializer.validated_data['shopfloor']
#             assembly_line = input_serializer.validated_data['assembly_line']
#             machine_id = input_serializer.validated_data['machine_id']
#             # start_shift = input_serializer.validated_data['start_shift']
#             # start_time_str = input_serializer.validated_data['start_time']
#             # # Convert start_time from string to datetime
#             # start_time = datetime.strptime(start_time_str, '%H:%M')
#         else:
#             return Response(input_serializer.errors, status=status.HTTP_400_BAD_REQUEST)


#         try:
#             product_recipe = models.productionPlanning.objects.get(job_id=job_id)
#             product_target = product_recipe.product_target
#         except models.productionPlanning.DoesNotExist:
#             product_target = None

#         if product_target:
#             product_target_seconds = product_target.total_seconds()
#         else:
#             # Set a default value or raise an exception as needed
#             product_target_seconds = 0


#         # Define the product target (60 seconds per unit)
#         product_target = product_target_seconds  # seconds

#         # Calculate total hours required to produce the quantity
#         total_seconds_required = quantity * product_target
#         total_hours_required = round(total_seconds_required / 3600, 2)  # 1 hour = 3600 seconds

#         # Calculate the number of days required
#         days_required = total_hours_required / 22  # Assuming 22 hours per day

#         # Initialize the current date and remaining hours
#         current_date = start_date
#         remaining_hours = total_hours_required
#         planned_hours = 0
#         total_planned_production = 0
#         balance_production = quantity

#         while remaining_hours > 0:
#             # Calculate planned production for the day
#             planned_hours = min(22, remaining_hours)

#             # Divide planned_hours by 2 for shift_a and shift_b
#             shift_a_planned_hours = shift_b_planned_hours = planned_hours // 2

            

#             planned_production = planned_hours * (3600 / product_target)
#             # Calculate balance production decrementally

#             total_planned_production += planned_production

#             balance_production = quantity - total_planned_production
            

#             # Create a LineMachineSlotConfig object and save it to the database
#             line_config = models.lineMachineSlotConfig(
#                 product_id=product_id,
#                 date=current_date,
#                 planned_hours=planned_hours,
#                 planned_production=planned_production,  # Planned production in units
#                 remaining_hours=remaining_hours - planned_hours,
#                 balance_production=balance_production,  # Balance production in units
#                 shift_a=f'08 - 20 ({shift_a_planned_hours})',
#                 shift_b=f'20 - 08 ({shift_b_planned_hours})',
#                 # shift_c=f'22 - 06 ({planned_hours})',
#                 shift_c= None,
#                 job_id=job_id,
#                 company=company,
#                 plant=plant,
#                 shopfloor=shopfloor,
#                 assembly_line=assembly_line,
#                 machine_id=machine_id,
#             )
#             line_config.save()

#             # Update current date and remaining hours for the next iteration
#             current_date += timedelta(days=1)
#             remaining_hours -= planned_hours

#         return Response({'message': 'Schedule calculated and saved.'}, status=status.HTTP_201_CREATED)
    

#     @action(detail=False, methods=['GET'])
#     def get_schedule(self, request):
#         # Add your code for handling GET requests here
#         # For example, you can retrieve and return schedule data
#         schedules = models.lineMachineSlotConfig.objects.all()
#         serializer = serializers.lineMachineSlotConfigSerializer(schedules, many=True)
#         return Response(serializer.data)
    
















'''Data Processing without consideration of start date and start time latest implementation of the above API'''
class LineMachineSlotConfigViewSet(viewsets.ModelViewSet):
    queryset = models.lineMachineSlotConfig.objects.all()
    serializer_class = serializers.lineMachineSlotConfigSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['id', 'job_id']

    @action(detail=False, methods=['POST'])
    def calculate_schedule(self, request):
        # Deserialize input data
        input_serializer = serializers.ScheduleInputSerializer(data=request.data)
        if input_serializer.is_valid():
            product_id = input_serializer.validated_data['product_id']
            quantity = input_serializer.validated_data['quantity']
            start_date = input_serializer.validated_data['start_date']
            job_id = input_serializer.validated_data['job_id']
            company = input_serializer.validated_data['company']
            plant = input_serializer.validated_data['plant']
            shopfloor = input_serializer.validated_data['shopfloor']
            assembly_line = input_serializer.validated_data['assembly_line']
            machine_id = input_serializer.validated_data['machine_id']
            # start_shift = input_serializer.validated_data['start_shift']
            # start_time_str = input_serializer.validated_data['start_time']
            # # Convert start_time from string to datetime
            # start_time = datetime.strptime(start_time_str, '%H:%M')

            try:
                # Check if a schedule already exists for the given machine_id and start_date
                existing_schedule = models.lineMachineSlotConfig.objects.get(
                    machine_id=machine_id,
                    date=start_date
                )

                # If a schedule already exists, return an error response
                return JsonResponse({'error': 'Schedule already exists for the given machine_id and start_date.'}, status=status.HTTP_400_BAD_REQUEST)

            except models.lineMachineSlotConfig.DoesNotExist:
                # Continue with the schedule calculation and creation
    
        


                try:
                    product_recipe = models.productionPlanning.objects.get(job_id=job_id)
                    product_target = product_recipe.product_target
                except models.productionPlanning.DoesNotExist:
                    product_target = None

                if product_target:
                    product_target_seconds = product_target.total_seconds()
                else:
                    # Set a default value or raise an exception as needed
                    product_target_seconds = 0


                # Define the product target (60 seconds per unit)
                product_target = product_target_seconds  # seconds

                # Calculate total hours required to produce the quantity
                total_seconds_required = quantity * product_target
                total_hours_required = round(total_seconds_required / 3600, 2)  # 1 hour = 3600 seconds

                # Calculate the number of days required
                days_required = total_hours_required / 22  # Assuming 22 hours per day

                # Initialize the current date and remaining hours
                current_date = start_date
                remaining_hours = total_hours_required
                planned_hours = 0
                total_planned_production = 0
                balance_production = quantity

                while remaining_hours > 0:
                    # Calculate planned production for the day
                    planned_hours = min(22, remaining_hours)

                    # Divide planned_hours by 2 for shift_a and shift_b
                    shift_a_planned_hours = shift_b_planned_hours = planned_hours // 2

                    

                    planned_production = planned_hours * (3600 / product_target)
                    # Calculate balance production decrementally

                    total_planned_production += planned_production

                    balance_production = quantity - total_planned_production
                    

                    # Create a LineMachineSlotConfig object and save it to the database
                    line_config = models.lineMachineSlotConfig(
                        product_id=product_id,
                        date=current_date,
                        planned_hours=planned_hours,
                        planned_production=planned_production,  # Planned production in units
                        remaining_hours=remaining_hours - planned_hours,
                        balance_production=balance_production,  # Balance production in units
                        shift_a=f'08 - 20 ({shift_a_planned_hours})',
                        shift_b=f'20 - 08 ({shift_b_planned_hours})',
                        # shift_c=f'22 - 06 ({planned_hours})',
                        shift_c= None,
                        job_id=job_id,
                        company=company,
                        plant=plant,
                        shopfloor=shopfloor,
                        assembly_line=assembly_line,
                        machine_id=machine_id,
                    )
                    line_config.save()

                    # Update current date and remaining hours for the next iteration
                    current_date += timedelta(days=1)
                    remaining_hours -= planned_hours

                return Response({'message': 'Schedule calculated and saved.'}, status=status.HTTP_201_CREATED)
    
        else:
            return Response(input_serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    

    @action(detail=False, methods=['GET'])
    def get_schedule(self, request):
        # Add your code for handling GET requests here
        # For example, you can retrieve and return schedule data
        schedules = models.lineMachineSlotConfig.objects.all()
        serializer = serializers.lineMachineSlotConfigSerializer(schedules, many=True)
        return Response(serializer.data)
    


class DateBlockAPIView(generics.ListAPIView):
    serializer_class = serializers.lineMachineSlotConfigSerializer  # Use your serializer class


    def get_queryset(self):
        # Calculate the first and last dates from the 'date' field
        first_date = models.lineMachineSlotConfig.objects.earliest('date').date
        last_date = models.lineMachineSlotConfig.objects.latest('date').date

        return [{'first_date': first_date, 'last_date': last_date}]

    def list(self, request, *args, **kwargs):
        queryset = self.get_queryset()
        return Response(queryset)






class LineMachineSlotConfigEdit(generics.RetrieveUpdateDestroyAPIView):
    queryset = models.lineMachineSlotConfig.objects.all()
    serializer_class = serializers.lineMachineSlotConfigSerializer
    lookup_url_kwarg = "id"



'''Get View for machineWiseDate Model'''
# class machineWiseDataView(generics.ListCreateAPIView):
#     queryset = models.machineWiseData.objects.order_by("time")
#     serializer_class = serializers.machineWiseDataSerializer
#     filter_backends = [DjangoFilterBackend]
#     filterset_fields = ['id', 'date', 'time', 'machine_id', 'product_target']

# '''Get View for machinewise group by machine id'''
# class machineWiseDataView(generics.ListAPIView):
#     # queryset = models.machineWiseData.objects.all()
#     serializer_class = serializers.machineWiseDataSerializer
    
#     now = datetime.now().time()
    
#     def get_queryset(self):
#         now = datetime.now().time()

#         if now >= datetime.strptime("08:00", "%H:%M").time() and now <= datetime.strptime("20:00", "%H:%M").time():
#             # Current time is between 08:00 and 20:00
#             queryset = models.machineWiseData.objects.filter(
#                 date=datetime.today(),
#                 time__in=[f"{i:02d}:00 - {i+1:02d}:00" for i in range(8, 20)]
#             )
#         else:
#             # Current time is between 20:00 and 08:00
#             today_midnight = datetime.combine(datetime.today(), datetime.min.time())
#             tomorrow_midnight = today_midnight + timedelta(days=1)

#             queryset = models.machineWiseData.objects.filter(
#                 date=datetime.today(),
#                 time__in=[f"{i:02d}:00 - {i+1:02d}:00" for i in range(20, 24)] +
#                          [f"{i:02d}:00 - {i+1:02d}:00" for i in range(0, 8)]
#             ).union(
#                 models.machineWiseData.objects.filter(
#                     date=tomorrow_midnight.date(),
#                     time__in=[f"{i:02d}:00 - {i+1:02d}:00" for i in range(0, 8)]
#                 )
#             )

#         return queryset.order_by('id')

import pytz

''' 2.0 Get View for machinewise group by machine id'''
class machineWiseDataView(generics.ListAPIView):
    serializer_class = serializers.machineWiseDataSerializer

    def get_queryset(self):
        ist_timezone = pytz.timezone('Asia/Kolkata')
        now = datetime.now(ist_timezone)

        if datetime.strptime("08:00", "%H:%M").time() <= now.time() <= datetime.strptime("20:00", "%H:%M").time():
            queryset = models.machineWiseData.objects.filter(
                date=now.date(),
                time__in=[f"{i:02d}:00 - {i+1:02d}:00" for i in range(8, 20)]
            )
        elif datetime.strptime("20:00", "%H:%M").time() <= now.time() <= datetime.strptime("23:59", "%H:%M").time():
            queryset = models.machineWiseData.objects.filter(
                date=now.date(),
                time__in=[f"{i:02d}:00 - {i+1:02d}:00" for i in range(20, 24)] + ["23:00 - 00:00"]
            )
        elif datetime.strptime("00:00", "%H:%M").time() <= now.time() <= datetime.strptime("08:00", "%H:%M").time():
            yesterday = now.date() - timedelta(days=1)
            queryset = models.machineWiseData.objects.filter(
                Q(date=yesterday, time__in=[f"{i:02d}:00 - {i+1:02d}:00" for i in range(20, 24)] + ["23:00 - 00:00"]) |
                Q(date=now.date(), time__in=[f"{i:02d}:00 - {i+1:02d}:00" for i in range(0, 8)])
            )
        else:
            # Handle unexpected cases or provide a default queryset
            # Example: return models.machineWiseData.objects.all()
            pass

        return queryset.order_by('id')




'''Update View for machineWiseDate Model'''
class machineWiseDataUpdate(generics.RetrieveUpdateDestroyAPIView):
    queryset = models.machineWiseData.objects.all()
    serializer_class = serializers.machineWiseDataUpdateSerializer
    lookup_url_kwarg = "id" 




class ProductionPlanningStatsView(generics.ListAPIView):
    queryset = models.productionPlanning.objects.all()
    serializer_class = serializers.productionPlanningSerializer

    def get_stats(self):
        today = date.today()
        planned_count = self.queryset.filter(planned_date__isnull=False, processing_date__isnull=True, completed_date__isnull=True).count()
        order_in_process_count = self.queryset.filter(processing_date__isnull=False, completed_date__isnull=True).count()
        completed_count = self.queryset.exclude(completed_date__isnull=True).count()
        transactions_today_count = self.queryset.filter(Q(assigned_date=today) | Q(planned_date=today) | Q(processing_date=today) | Q(completed_date=today)).count()

        return {
            'planned': planned_count,
            'order_in_process': order_in_process_count,
            'completed': completed_count,
            'transactions_today': transactions_today_count,
        }

    def list(self, request, *args, **kwargs):
        stats = self.get_stats()
        return Response(stats)
    

class lineMachineSlotConfigViewAll(generics.ListAPIView):
    queryset = models.lineMachineSlotConfig.objects.all()
    serializer_class = serializers.lineMachineSlotConfigSerializer



class ExportExcelView(View):
    def get(self, request, *args, **kwargs):
        queryset = self.get_filtered_queryset(request)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Machine Records"

        headers = ["Employee ID", "Employee Name", "Log Date", "Company", "Shift Status"]
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num, value=header)

        for row_num, record in enumerate(queryset, 2):
            ws.cell(row=row_num, column=1, value=record.employeeid)
            ws.cell(row=row_num, column=2, value=record.employee_name)
            ws.cell(row=row_num, column=3, value=record.logdate)
            ws.cell(row=row_num, column=4, value=record.company)
            ws.cell(row=row_num, column=5, value=record.shift_status)

        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = "attachment; filename=attendance_records.xlsx"
        wb.save(response)

        return response

    def get_filtered_queryset(self, request):
        queryset = super().get_queryset()
        search_query = self.request.query_params.get('search', None)
        date_query = self.request.query_params.get('date', None)  

        if date_query:
            try:
                parsed_date = self.parse_date_with_format(date_query, "%Y/%m/%d")  
                next_day = parsed_date + timedelta(days=1)  

                queryset = queryset.filter(
                    logdate__gte=parsed_date,
                    logdate__lt=next_day
                )
            except Exception as e:
                print("Error parsing date:", e)

        if search_query:
            queryset = queryset.filter(
                Q(employee_name__icontains=search_query) |
                Q(employeeid__icontains=search_query)
            )
        return queryset
    

class AssemblyLineWiseDataView(generics.ListAPIView):
    queryset = models.assemblyLineWiseData.objects.all()
    serializer_class = serializers.AssemblyLineWiseDataSerializer

    def list(self, request, *args, **kwargs):
        last_object = self.get_queryset().last()
        serializer = self.get_serializer([last_object], many=True)
        return Response(serializer.data)
    

class AssemblyLineWiseDataUpdate(generics.RetrieveUpdateDestroyAPIView):
    queryset = models.assemblyLineWiseData.objects.all()
    serializer_class = serializers.AssemblyLineWiseDataSerializer
    lookup_url_kwarg = "id"

class soloAssemblyLineDataView(generics.ListAPIView):
    # queryset = models.soloAssemblyLineData.objects.all()
    serializer_class = serializers.soloAssemblyLineDataSerializer

    def get_queryset(self):
        current_datetime = datetime.now()
        current_date = current_datetime.date()
        current_time = current_datetime.time()

        # Determine the shift based on the current time
        shift = 'FS' if time(7, 0) <= current_time <= time(19, 0) else 'SS'

        # Filter the queryset based on the current date and shift
        queryset = models.soloAssemblyLineData.objects.filter(date=current_date, shift=shift)

        # Arrange the queryset by 'stage_no' in ascending order
        queryset = queryset.order_by('stage_no')

        print("Current Date:", current_date)
        print("Shift:", shift)

        return queryset

class soloAssemblyLineDataUpdate(generics.RetrieveUpdateDestroyAPIView):
    queryset = models.soloAssemblyLineData.objects.all()
    serializer_class = serializers.soloAssemblyLineDataUpdateSerializer
    lookup_url_kwarg = "id"



class spellAssemblyLineDataView(generics.ListAPIView):
    # queryset = models.soloAssemblyLineData.objects.all()
    serializer_class = serializers.spellAssemblyLineDataSerializer

    def get_queryset(self):
        current_datetime = datetime.now()
        current_date = current_datetime.date()
        current_time = current_datetime.time()

        # Determine the shift based on the current time
        shift = 'FS' if time(7, 0) <= current_time <= time(19, 0) else 'SS'

        # Filter the queryset based on the current date and shift
        queryset = models.spellAssemblyLineData.objects.filter(date=current_date, shift=shift)

        # Arrange the queryset by 'stage_no' in ascending order
        queryset = queryset.order_by('stage_no')

        print("Current Date:", current_date)
        print("Shift:", shift)

        return queryset

class spellAssemblyLineDataUpdate(generics.RetrieveUpdateDestroyAPIView):
    queryset = models.spellAssemblyLineData.objects.all()
    serializer_class = serializers.spellAssemblyLineDataUpdateSerializer
    lookup_url_kwarg = "id"
    







class ProductionAndonView(generics.ListAPIView):

    # queryset = models.ProductionAndon.objects.order_by('-machine_datetime')[:1]
    serializer_class = serializers.ProductionAndonSerializer

    def get_queryset(self):
        # Get the current date and time in the Asia/Kolkata timezone
        today = datetime.now(pytz.timezone('Asia/Kolkata'))

        # Retrieve the latest object from the queryset
        latest_object = models.ProductionAndon.objects.order_by('-machine_datetime').first()

        default_value = [{"r": None }]
        # default_value = models.ProductionAndon.objects.none()

        # Check if the latest object exists
        if latest_object:
            # Get the datetime of the latest object
            latest_datetime = latest_object.machine_datetime

            # Compare today's date and time with the datetime of the latest object
            if today.date() == latest_datetime.date():
                print("Today's date matches the date of the latest object.")
            else:
                print("Today's date does not match the date of the latest object.")

            # Define tolerance in minutes
            tolerance_minutes = 2

            # Calculate the time range within the tolerance
            lower_bound = latest_datetime - timedelta(minutes=tolerance_minutes)
            upper_bound = latest_datetime + timedelta(minutes=tolerance_minutes)

            # Check if the current time falls within the tolerance range
            if lower_bound.time() <= today.time() <= upper_bound.time():
                today_time = today.time() 
                print(today_time)
                print("The current time is within {} minutes of the time of the latest object.".format(tolerance_minutes))
                # Return the queryset
                return models.ProductionAndon.objects.order_by('-machine_datetime')[:1]
            
            else:
                print("The current time is not within {} minutes of the time of the latest object.".format(tolerance_minutes))
                # return default_value
                return [{"error": "No objects found in the queryset."}]

        else:
            print("No objects found in the queryset.")
            # return default_value
            return [{"error": "No objects found in the queryset."}]

        # return models.ProductionAndon.objects.order_by('-machine_datetime')[:1]




class ExcelImportListView(generics.ListAPIView):
    queryset = models.ExcelImport.objects.all()
    serializer_class = serializers.ExcelImportSerializer



class ExcelImportView(APIView):
    def post(self, request, *args, **kwargs):
        # Assuming the file is sent in the 'file' field of the request
        file = request.FILES.get('file')

        if not file:
            return Response({'error': 'No file uploaded'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            workbook = openpyxl.load_workbook(file, read_only=True)
            sheet = workbook.active

            # Assuming your Excel file has headers and data starts from the second row
            headers = [cell.value for cell in sheet[1]]
            data = []

            for row in sheet.iter_rows(min_row=2, values_only=True):
                row_data = dict(zip(headers, row))
                data.append(row_data)

            serializer = serializers.ExcelImportSerializer(data=data, many=True)

            if serializer.is_valid():
                serializer.save()
                return Response({'success': 'Data imported successfully'}, status=status.HTTP_201_CREATED)
            else:
                return Response({'error': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)

        except Exception as e:
            return Response({'error': f'Error processing Excel file: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        


class ExportExcelTemplate(View):
    def get(self, request, *args, **kwargs):
        queryset = models.ExcelImport.objects.all()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Template"

        headers = [
            "field1", "field2"
        ]

        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num, value=header)

        for row_num, record in enumerate(queryset, 2):
            ws.cell(row=row_num, column=1, value=record.field1)
            ws.cell(row=row_num, column=2, value=record.field2)

        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = "attachment; filename=machine_records.xlsx"
        wb.save(response)

        return response
    

from django.db.models import Max
from django.db.models.functions import TruncHour

class HourlyProductionAndon(APIView):
    def get(self, request, *args, **kwargs):
        try:
            # Get the latest 'p' value at the end of each hour
            hourly_data = (
                models.ProductionAndon.objects
                .annotate(hour=TruncHour('machine_datetime'))
                .values('hour')
                .annotate(last_p=Max('p'))
            )

            # Extracting the 'p' value at the end of each hour
            result = [{'hour': entry['hour'], 'last_p': entry['last_p']} for entry in hourly_data]

            return Response(result, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        


class ExportExcelMachineView(View):
    def get(self, request, *args, **kwargs):
        # queryset = models.machineWiseData.objects.all()
        queryset = models.machineWiseData.objects.all().order_by('date', 'time')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Machine Records"

        headers = [
            "Plant", "Shopfloor", "Assembly Line", "Machine ID", "Product ID", "Lot/Batch",
            "Product Target", "Shift", "Date", "Time", "Batch Order Quantity", "Order Processed", "Order Pending", "On Time", "Idle Time", "Idle Reason", "Break",
            "Actual", "Target", "Performance", "Gap", "kW-h"
        ]

        # Set font style and background color for headers
        header_font = Font(size=14, bold=True)
        header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

        footer_font = Font(bold=True)
        footer_fill = PatternFill(start_color="799184", end_color="799184", fill_type="solid")

        # for col_num, header in enumerate(headers, 1):
        #     cell = ws.cell(row=1, column=col_num, value=header)
        #     cell.font = header_font
        #     cell.fill = header_fill

        # Retrieve the last record from productionPlanning
        last_record = models.productionPlanning.objects.latest('id')

        # Get the quantity from the last record
        quantity = last_record.quantity

        processed_quantity = 0
        remaining_quantity = quantity

        row_num = 1
        current_shift = None
        
        shift_total_on_time = 0
        shift_total_idle_time = 0
        shift_total_actual = 0
        shift_total_target = 0
        shift_total_performance = 0
        performance_count = 0
        shift_average_performance = 0
        shift_total_gap = 0

        shift_total_order_quantity = quantity

        for record in queryset:
            # Extract the hour from the record.time in 24-hour format
            record_hour = int(record.time.split(':')[0])

            # Determine the shift based on the time
            if 8 <= record_hour < 20:
                shift = "Shift A, 08 - 20 (11)"
            else:
                shift = "Shift B, 20 - 08 (11)"

            if shift != current_shift:
                # New shift detected, write the previous shift's total target
                if current_shift:
                    ws.cell(row=row_num, column=1, value=f"{current_shift} => Total")
                    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=10)  # Merge columns 1 to 4
                    for col_num in range(1, 23):  # Loop through columns 1 to 19
                        cell = ws.cell(row=row_num, column=col_num)
                        cell.font = footer_font
                        cell.fill = footer_fill
                    ws.cell(row=row_num, column=11, value=shift_total_order_quantity)
                    ws.cell(row=row_num, column=12, value=processed_quantity)
                    ws.cell(row=row_num, column=13, value=remaining_quantity)
                    ws.cell(row=row_num, column=14, value=shift_total_on_time)
                    ws.cell(row=row_num, column=15, value=shift_total_idle_time)
                    ws.cell(row=row_num, column=18, value=shift_total_actual)
                    ws.cell(row=row_num, column=19, value=shift_total_target)
                    ws.cell(row=row_num, column=20, value=shift_average_performance)
                    ws.cell(row=row_num, column=21, value=shift_total_gap)
                    ws.cell(row=row_num, column=22, value="")
                    row_num += 1

                    ws.cell(row=row_num, column=1, value="")
                    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=22)
                    row_num += 1

                # Reset the shift target sum for the new shift
                shift_total_on_time = 0
                shift_total_idle_time = 0
                shift_total_actual = 0
                shift_total_target = 0
                shift_total_performance = 0
                performance_count = 0
                shift_average_performance = 0
                shift_total_gap = 0


                current_shift = shift
            
                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=row_num, column=col_num, value=header)
                    cell.font = header_font
                    cell.fill = header_fill

                     # Calculate the width of the column based on the length of the header
                    header_length = len(header)
                    column_width = max(16, header_length + 2)  # Minimum width of 10, adjust as needed
                    ws.column_dimensions[get_column_letter(col_num)].width = column_width
                    
                row_num += 1

            # processed_quantity += record.actual

            # Write the record data
            ws.cell(row=row_num, column=1, value=record.plant)
            ws.cell(row=row_num, column=2, value=record.shopfloor)
            ws.cell(row=row_num, column=3, value=record.assembly_line)
            ws.cell(row=row_num, column=4, value=record.machine_id)
            ws.cell(row=row_num, column=5, value="AQUA 1000ml")
            ws.cell(row=row_num, column=6, value="")
            ws.cell(row=row_num, column=7, value=record.product_target)
            ws.cell(row=row_num, column=8, value=shift)
            ws.cell(row=row_num, column=9, value=record.date)
            ws.cell(row=row_num, column=10, value=record.time)
            ws.cell(row=row_num, column=11, value=quantity)
            ws.cell(row=row_num, column=12, value=processed_quantity)
            ws.cell(row=row_num, column=13, value=remaining_quantity)
            ws.cell(row=row_num, column=14, value=record.on_time)
            ws.cell(row=row_num, column=15, value=record.idle_time)
            ws.cell(row=row_num, column=16, value="")
            ws.cell(row=row_num, column=17, value="")
            ws.cell(row=row_num, column=18, value=record.actual)
            ws.cell(row=row_num, column=19, value=record.target)
            ws.cell(row=row_num, column=20, value=record.performance)
            ws.cell(row=row_num, column=21, value=record.gap)
            ws.cell(row=row_num, column=22, value=record.current)

            if record.on_time is not None:
                shift_total_on_time += record.on_time
            # shift_total_on_time += record.on_time
            if record.idle_time is not None:
                shift_total_idle_time += record.idle_time
            if record.actual is not None:
                shift_total_actual += record.actual
            # shift_total_actual += record.actual
            if record.target is not None:
                shift_total_target += record.target
            # shift_total_target += record.target
            performance_count += 1
            shift_average_performance = shift_total_performance / performance_count if performance_count > 0 else 0
            shift_total_gap = shift_total_actual - shift_total_target

            if record.actual is not None:
                processed_quantity += record.actual
            if record.actual is not None:
                remaining_quantity -= record.actual

            row_num += 1

        # Write the total target for the last shift
        ws.cell(row=row_num, column=1, value=f"{current_shift} => Total")
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=10)  # Merge columns 1 to 4
        for col_num in range(1, 23):  # Loop through columns 1 to 19
            cell = ws.cell(row=row_num, column=col_num)
            cell.font = footer_font
            cell.fill = footer_fill
        ws.cell(row=row_num, column=11, value=shift_total_order_quantity)
        ws.cell(row=row_num, column=12, value=processed_quantity)
        ws.cell(row=row_num, column=13, value=remaining_quantity)
        ws.cell(row=row_num, column=14, value=shift_total_on_time)
        ws.cell(row=row_num, column=15, value=shift_total_idle_time)
        ws.cell(row=row_num, column=18, value=shift_total_actual)
        ws.cell(row=row_num, column=19, value=shift_total_target)
        ws.cell(row=row_num, column=20, value=shift_average_performance)
        ws.cell(row=row_num, column=21, value=shift_total_gap)
        ws.cell(row=row_num, column=22, value="")

        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = "attachment; filename=machine_records.xlsx"
        wb.save(response)

        return response
    

class ExportProductionInfoExcel(View):
    def get(self, request, *args, **kwargs):
        # queryset = models.machineWiseData.objects.all()
        queryset = models.productionPlanning.objects.all()

        lmc = models.lineMachineConfig.objects.all()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Machine Records"

        headers = [
            "Jobwork", "Job Assigned", "Customer", "Po & Date", "Drawing No", "Plant", "Shopfloor", "Assembly Line", 
            "Machine ID", "Product ID", "Lot/Batch", "Product Target", "Shift", "Date", "Time", "Batch Order Quantity",
            "Order Processed", "Order Pending", "On Time", "Idle Time", "Idle Reason", "Break", "Actual", "Target",
            "Performance", "Gap", "kW-h"
        ]

        row_num = 1

        # Set font style and background color for headers
        header_font = Font(size=14, bold=True)
        header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=row_num, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill

            # Calculate the width of the column based on the length of the header
            header_length = len(header)
            column_width = max(20, header_length + 2)  # Minimum width of 20, adjust as needed
            ws.column_dimensions[get_column_letter(col_num)].width = column_width
                    
        row_num += 1

        
        # Dictionary to store plant, shopfloor, and assemblyline details for each job_id
        job_details = {}

        # Populate job_details dictionary with lineMachineConfig details
        for obj2 in lmc:
            job_id = obj2.job_id
            if job_id not in job_details:
                job_details[job_id] = {}
            job_details[job_id]['plant'] = obj2.plant
            job_details[job_id]['shopfloor'] = obj2.shopfloor
            job_details[job_id]['assemblyline'] = obj2.assembly_line
            job_details[job_id]['machine_id'] = obj2.machine_id
            job_details[job_id]['product_id'] = obj2.product_id

        # ws.cell(row=row_num, column=1, value=models.productionPlanning.job_id)
        for obj in queryset:

            job_id = obj.job_id
            plant = job_details[job_id]['plant'] if job_id in job_details else ""
            shopfloor = job_details[job_id]['shopfloor'] if job_id in job_details else ""
            assemblyline = job_details[job_id]['assemblyline'] if job_id in job_details else ""
            machine_id = job_details[job_id]['machine_id'] if job_id in job_details else ""
            product_id = job_details[job_id]['product_id'] if job_id in job_details else ""

            ws.append([
                obj.job_id,
                obj.assigned_date,
                obj.customer,
                obj.po_no,
                obj.drawing_no,
                plant,
                shopfloor,
                assemblyline,
                machine_id,
                product_id,
                obj.batch_no,
            ])


        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = "attachment; filename=production_info.xlsx"
        wb.save(response)

        return response
